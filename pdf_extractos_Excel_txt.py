import fitz  # PyMuPDF
import pandas as pd
import re
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, Alignment

# --- MODO DE DEPURACIÓN ---
# Cambia a False para desactivar los mensajes de diagnóstico en la consola
DEBUG_MODE = False

# extract_transactions_from_pdf function (as previously modified, remains unchanged)
def extract_transactions_from_pdf(pdf_path):
    """
    Extracts credit card transaction data for multiple salespeople from a PDF,
    first into a dictionary keyed by salesperson, then into a DataFrame.
    Includes total rows and blank lines as specified. Handles page breaks
    and associates transactions with the correct salesperson.
    """
    doc = fitz.open(pdf_path)
    salesperson_data = {}
    
    start_extraction = False
    # --- Variables de estado simplificadas ---
    current_salesperson = None  # La ÚNICA variable que controla el estado principal
    current_headers_coords = {}
    col_mapping_order = []
    # -----------------------------------------
    latest_date = None

    output_columns = ["FECHA", "DESCRIPCIÓN", "NRO. CUPÓN", "PESOS", "DÓLARES"]

    consumos_pattern = re.compile(r"Consumos\s+([A-Z\s.]+)", re.IGNORECASE)
    total_consumos_pattern = re.compile(r"TOTAL CONSUMOS DE\s+([A-Z\s.]+)", re.IGNORECASE)

    for page_num in range(doc.page_count):
        # --- DEBUG: INICIO DE PÁGINA ---
        if DEBUG_MODE:
            print("\n" + "="*50)
            print(f"--- PROCESANDO PÁGINA {page_num + 1} ---")
            print(f"[ESTADO INICIAL] Vendedor actual: '{current_salesperson}'. Coordenadas de encabezado presentes: {bool(current_headers_coords)}")
            print("="*50)
        # --- FIN DEBUG ---
        page = doc.load_page(page_num)
        text_blocks = page.get_text("blocks")

        if not start_extraction:
            for block in text_blocks:
                if "DETALLE" in block[4]:
                    start_extraction = True
                    if DEBUG_MODE: print("[INFO] 'DETALLE' encontrado. Iniciando extracción general.")
                    break
            if not start_extraction:
                continue

        for i, block in enumerate(text_blocks):
            block_text = block[4].strip()

            # --- Lógica de PARADA (sin cambios) ---
            if re.search(r"Impuestos,\s*cargos\s*e\s*intereses", block_text, re.IGNORECASE) or \
               re.search(r"Legales\s*y\s*avisos", block_text, re.IGNORECASE) or \
               (re.search(r"Tarjetas\s*de\s*Crédito", block_text, re.IGNORECASE) and block[1] < 200):
                if DEBUG_MODE: print(f"[INFO] Condición de parada encontrada en bloque: '{block_text[:50]}...'. Finalizando extracción.")
                start_extraction = False
                current_salesperson = None
                current_headers_coords = {}
                col_mapping_order = []
                break

            is_header_block = "FECHA" in block_text.upper() and ("PESOS" in block_text.upper() or "DÓLARES" in block_text.upper())

            if is_header_block and current_salesperson:
                if DEBUG_MODE: print(f"[INFO] Analizando bloque que podría ser un encabezado para '{current_salesperson}': \"{block_text.replace(chr(10), ' ')}\"")
                header_block_coords = block[:4]
                words_in_header_block = page.get_text("words", clip=header_block_coords)
                words_in_header_block.sort(key=lambda x: x[0])
                is_cristian_palet = "CRISTIAN A PALET" in current_salesperson.upper()
                target_headers = ["FECHA", "DESCRIPCIÓN", "NRO. CUPÓN"]
                if is_cristian_palet and "DÓLARES" in block_text.upper() and block_text.upper().find("DÓLARES") < block_text.upper().find("PESOS"):
                    target_headers.extend(["DÓLARES", "PESOS"])
                else:
                    target_headers.extend(["PESOS", "DÓLARES"])
                
                y_tolerance = 5
                candidate_header_words = [w for w in words_in_header_block if abs(w[1] - header_block_coords[1]) < y_tolerance]
                header_x0s = {}
                for target_header in target_headers:
                    found_x0 = None
                    target_key = target_header.replace(' ', '').replace('.', '').upper()
                    for word_info in candidate_header_words:
                        if target_key in word_info[4].replace(' ', '').replace('.', '').upper():
                            found_x0 = word_info[0]
                            break
                    if found_x0 is not None:
                        header_x0s[target_header] = found_x0
                    else:
                        if target_header == "FECHA": header_x0s["FECHA"] = 36
                        elif target_header == "DESCRIPCIÓN": header_x0s["DESCRIPCIÓN"] = 95
                        elif target_header == "NRO. CUPÓN": header_x0s["NRO. CUPÓN"] = 300
                        elif target_header == "PESOS": header_x0s["PESOS"] = 390
                        elif target_header == "DÓLARES": header_x0s["DÓLARES"] = 480

                TOLERANCIA = 6  # Margen de error en píxeles.

                if DEBUG_MODE: print(f"[DEBUG] Coordenadas de encabezado originales: {header_x0s}")

                # --- SOLUCIÓN: APLICAR MARGEN DE ERROR A TODAS LAS COLUMNAS ---
                for col_name in header_x0s:
                    original_x0 = header_x0s[col_name]
                    header_x0s[col_name] = max(0, original_x0 - TOLERANCIA)
                    if DEBUG_MODE:
                        print(f"[FIX] Aplicada tolerancia de {TOLERANCIA}px a la columna '{col_name}'. Coordenada de inicio: {header_x0s[col_name]}")
                
                if DEBUG_MODE: print(f"[DEBUG] Coordenadas de encabezado finales con tolerancia: {header_x0s}")

                current_headers_coords = {}
                col_mapping_order = []
                for k_idx, col_name in enumerate(target_headers):
                    x0 = header_x0s.get(col_name, 0)
                    x1 = 1000

                    if col_name == "FECHA":
                        x1 = 95
                    elif k_idx + 1 < len(target_headers):
                        next_col_name = target_headers[k_idx + 1]
                        if header_x0s.get(next_col_name):
                            x1 = header_x0s.get(next_col_name)
                        else:
                            x1 = x0 + 100 
                    
                    current_headers_coords[col_name] = (x0, x1)
                    col_mapping_order.append(col_name)
                
                if DEBUG_MODE:
                    print(f"[INFO] ¡Encabezado confirmado y procesado! Orden de columnas: {col_mapping_order}")
                    print(f"[DEBUG] Coordenadas de columna finales: {current_headers_coords}")

            consumos_match = consumos_pattern.match(block_text)
            if consumos_match:
                current_salesperson = consumos_match.group(1).strip()
                if current_salesperson not in salesperson_data:
                    salesperson_data[current_salesperson] = []
                if DEBUG_MODE: print(f"\n[INFO] CAMBIO DE CONTEXTO: Nuevo vendedor encontrado -> '{current_salesperson}'")
                
                current_headers_coords = {}
                col_mapping_order = []
                if DEBUG_MODE: print("[INFO] Coordenadas de encabezado reseteadas. Esperando nuevo encabezado.")
                continue

            total_consumos_match = total_consumos_pattern.match(block_text.upper())
            if total_consumos_match and current_salesperson:
                if DEBUG_MODE: print(f"\n[INFO] CAMBIO DE CONTEXTO: Total encontrado para -> '{current_salesperson}'")
                total_row_data = {col: "" for col in output_columns}
                total_row_data["DESCRIPCIÓN"] = block_text.strip()
                total_row_data["_TYPE"] = "TOTAL"
                total_row_data["NAME"] = current_salesperson
                amounts = re.findall(r"([-+]?\d{1,3}(?:\.\d{3})*(?:,\d+)?|\d+,\d+)", block_text)
                if len(amounts) >= 2:
                    pesos_str_raw, dolares_str_raw = amounts[-2], amounts[-1]
                    cleaned_pesos = pesos_str_raw.replace('.', '').replace(',', '.')
                    if cleaned_pesos.endswith('-'): cleaned_pesos = '-' + cleaned_pesos[:-1]
                    try: total_row_data["PESOS"] = float(cleaned_pesos)
                    except ValueError: total_row_data["PESOS"] = ""
                    cleaned_dolares = dolares_str_raw.replace('.', '').replace(',', '.')
                    if cleaned_dolares.endswith('-'): cleaned_dolares = '-' + cleaned_dolares[:-1]
                    try: total_row_data["DÓLARES"] = float(cleaned_dolares)
                    except ValueError: total_row_data["DÓLARES"] = ""
                if current_salesperson in salesperson_data:
                     salesperson_data[current_salesperson].append(total_row_data)
                
                current_salesperson = None
                current_headers_coords = {}
                col_mapping_order = []
                if DEBUG_MODE: print("[INFO] Estado y coordenadas reseteados después del total.")
                continue

            if current_salesperson and current_headers_coords:
                current_block_words = page.get_text("words", clip=block[:4])
                lines_grouped_by_y = {}
                y_tolerance_lines = 5
                for word_info in current_block_words:
                    word_y0 = word_info[1]
                    found_line = False
                    for line_y_key in lines_grouped_by_y.keys():
                        if abs(word_y0 - line_y_key) < y_tolerance_lines:
                            lines_grouped_by_y[line_y_key].append(word_info)
                            found_line = True
                            break
                    if not found_line:
                        lines_grouped_by_y[word_y0] = [word_info]
                
                for line_y in sorted(lines_grouped_by_y.keys()):
                    words_on_current_line = sorted(lines_grouped_by_y[line_y], key=lambda x: x[0])
                    
                    if DEBUG_MODE:
                        line_text_for_debug = ' '.join([w[4] for w in words_on_current_line])
                        print(f"\n----------------------------------------------------")
                        print(f"[DEBUG] Procesando línea de texto: {line_text_for_debug.split()}")

                    row_data = {col: "" for col in output_columns}
                    temp_col_values = {col: [] for col in current_headers_coords.keys()}

                    for word_info in words_on_current_line:
                        word_text, word_x0 = word_info[4], word_info[0]
                        for col_name, (x0_col, x1_col) in current_headers_coords.items():
                            if x0_col <= word_x0 < x1_col:
                                temp_col_values[col_name].append(word_text)
                                break
                    
                    for col_name in col_mapping_order:
                        extracted_value = " ".join(temp_col_values.get(col_name, [])).strip()
                        if col_name in row_data:
                            row_data[col_name] = extracted_value.replace('--', '-').replace(',,', ',').replace('. .', '.')
                    
                    nro_cupon_val = row_data.get("NRO. CUPÓN", "")
                    if ' ' in nro_cupon_val and not row_data.get("PESOS"):
                        parts = nro_cupon_val.split(' ', 1)
                        if re.search(r'[\d,.-]+', parts[1]):
                            row_data["NRO. CUPÓN"], row_data["PESOS"] = parts[0], parts[1]
                    
                    if DEBUG_MODE: print(f"[DEBUG] Fila construida: {row_data}")
                    
                    date_match = re.match(r"^\d{1,2}[-/\s]?(?:Jan|Ene|Feb|Mar|Abr|Apr|May|Jun|Jul|Ago|Sep|Oct|Nov|Dic)[-/\s]?\d{2}$", row_data["FECHA"], re.IGNORECASE)
                    
                    if not date_match:
                        if DEBUG_MODE: print(f"[DEBUG] Resultado del match de fecha para '{row_data['FECHA']}': RECHAZADO")
                        continue
                    
                    if DEBUG_MODE: print(f"[DEBUG] Resultado del match de fecha para '{row_data['FECHA']}': ACEPTADO")
                    if DEBUG_MODE: print(f"[SUCCESS] Transacción guardada para {current_salesperson}: {row_data['FECHA']} - {row_data['DESCRIPCIÓN']}")

                    month_map = {'ene': 'Jan', 'abr': 'Apr', 'ago': 'Aug', 'dic': 'Dec'}
                    date_str = date_match.group(0).lower()
                    for spa, eng in month_map.items():
                        date_str = date_str.replace(spa, eng)

                    try:
                        date_str_normalized = date_str.replace(' ', '-').replace('--', '-')
                        date_parts = date_str_normalized.split('-')
                        if len(date_parts[-1]) == 2:
                            full_year = (datetime.now().year // 100) * 100 + int(date_parts[-1])
                            date_obj = datetime.strptime(f"{date_parts[0]}-{date_parts[1]}-{full_year}", "%d-%b-%Y")
                        else:
                            date_obj = datetime.strptime(date_str_normalized, "%d-%b-%Y")
                        if latest_date is None or date_obj > latest_date:
                            latest_date = date_obj
                    except (ValueError, IndexError):
                        pass

                    for col in ["PESOS", "DÓLARES"]:
                        if row_data[col]:
                            amount_str = str(row_data[col]).replace('.', '').replace(',', '.')
                            if amount_str.endswith('-'): amount_str = '-' + amount_str[:-1]
                            try: row_data[col] = float(amount_str)
                            except ValueError: row_data[col] = ""
                    
                    row_data["_TYPE"] = "TRANSACTION"
                    row_data["NAME"] = current_salesperson
                    if current_salesperson in salesperson_data:
                        salesperson_data[current_salesperson].append(row_data)

    doc.close()

    final_data_for_df = []
    for name, items_list in salesperson_data.items():
        final_data_for_df.append({col: f"--- Consumos {name} ---" if col == "DESCRIPCIÓN" else "" for col in output_columns})
        final_data_for_df.append({col: col for col in output_columns})
        for item_dict in items_list:
            final_data_for_df.append({col: item_dict.get(col, "") for col in output_columns})
            if item_dict.get("_TYPE") == "TOTAL":
                final_data_for_df.append({col: "" for col in output_columns})
    
    return pd.DataFrame(final_data_for_df, columns=output_columns), latest_date

def save_to_excel(dataframe, latest_date, output_folder="output_excel", filename="transactions.xlsx"):
    """
    Saves a Pandas DataFrame to an Excel file in a specified folder,
    applying basic formatting.

    Args:
        dataframe (pandas.DataFrame): The DataFrame to save.
        latest_date (datetime): The latest date for the title.
        output_folder (str): The name of the folder to save the Excel file in.
        filename (str): The name of the Excel file.
    """
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        print(f"Created folder: {output_folder}")

    output_path = os.path.join(output_folder, filename)

    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    dataframe.to_excel(writer, index=False, sheet_name="Transactions", startrow=1, header=False)
    
    workbook = writer.book
    sheet = writer.sheets["Transactions"]

    df_columns = list(dataframe.columns)

    if latest_date:
        title_month = latest_date.strftime("%B")
        title_text = f"Monthly Consumption {title_month}"
        sheet['A1'] = title_text
        sheet['A1'].font = Font(bold=True, size=14)
        sheet.merge_cells('A1:E1')
        title_cell = sheet['A1']
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

    for col_idx, col_name_str in enumerate(df_columns):
        max_length = 0
        column_letter = get_column_letter(col_idx + 1)
        column = sheet.column_dimensions[column_letter]
        
        header_length = len(col_name_str)
        max_length = max(max_length, header_length)
        for cell in sheet[column_letter]:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if isinstance(cell.value, (int, float)):
                        cell_length = len(f"{cell.value:,.2f}")
                    max_length = max(max_length, cell_length)
            except:
                pass
        
        adjusted_width = (max_length + 2)

        if col_name_str in ["PESOS", "DÓLARES"]:
            column.width = 20
        else:
            column.width = adjusted_width
    
    try:
        pesos_col_letter_idx = df_columns.index("PESOS") + 1
        dollars_col_letter_idx = df_columns.index("DÓLARES") + 1
        description_col_letter_idx = df_columns.index("DESCRIPCIÓN") + 1
        fecha_col_letter_idx = df_columns.index("FECHA") + 1
    except ValueError:
        print("Error: One of the expected columns is not in the DataFrame.")
        writer.close()
        return

    custom_currency_format = '_-$ * #,##0.00_-;-$ * #,##0.00_-;_-$ * "-"??_-;_-@_-'

    for row_idx in range(2, sheet.max_row + 1):
        pesos_cell = sheet.cell(row=row_idx, column=pesos_col_letter_idx)
        if isinstance(pesos_cell.value, (int, float)):
            pesos_cell.number_format = custom_currency_format

        dollars_cell = sheet.cell(row=row_idx, column=dollars_col_letter_idx)
        if isinstance(dollars_cell.value, (int, float)):
            dollars_cell.number_format = custom_currency_format

        description_cell = sheet.cell(row=row_idx, column=description_col_letter_idx)
        if description_cell.value and "TOTAL CONSUMOS DE" in str(description_cell.value).upper():
            match = re.search(r"(TOTAL CONSUMOS DE\s+[A-Z\s.]+)", str(description_cell.value).upper())
            if match:
                description_cell.value = match.group(1).strip()

    thin_dotted_border = Border(left=Side(style='dotted'),
                                right=Side(style='dotted'),
                                top=Side(style='dotted'),
                                bottom=Side(style='dotted'))

    for row_idx in range(2, sheet.max_row + 1):
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            is_heading_row = False
            description_cell_value = sheet.cell(row=row_idx, column=description_col_letter_idx).value
            if description_cell_value and (str(description_cell_value).startswith("TOTAL CONSUMOS DE") or
                                          str(description_cell_value).startswith("DESCRIPCIÓN")):
                is_heading_row = True
            
            is_blank_row = not any(sheet.cell(row=row_idx, column=c_idx).value for c_idx in range(1, sheet.max_column + 1))

            if is_heading_row and not is_blank_row:
                cell.border = thin_dotted_border
    
    writer.close()
    print(f"Data successfully saved to {output_path}")


# <<< NUEVO >>>: Función para guardar los datos en un archivo de texto
def save_to_txt(dataframe, latest_date, output_folder="output_txt", filename="consumos.txt"):
    """
    Guarda un DataFrame en un archivo de texto con formato de ancho fijo.
    Aplica una limpieza a las filas de totales para un formato correcto.

    Args:
        dataframe (pandas.DataFrame): El DataFrame a guardar.
        latest_date (datetime): La fecha más reciente para el título.
        output_folder (str): La carpeta donde se guardará el archivo.
        filename (str): El nombre del archivo de texto.
    """
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        print(f"Created folder: {output_folder}")

    output_path = os.path.join(output_folder, filename)

    # <<< CORRECCIÓN >>>: Se crea una copia del DataFrame para limpiarlo sin afectar el original (que usa el Excel).
    df_txt = dataframe.copy()

    # <<< CORRECCIÓN >>>: Se define una función para limpiar solo el texto de la descripción en las filas de totales.
    # Esta lógica es similar a la que ya se usaba en la función de Excel para garantizar consistencia.
    total_pattern = re.compile(r"(TOTAL CONSUMOS DE\s+[A-Z\s\.]+)", re.IGNORECASE)
    
    def clean_total_description(description):
        # Revisa si el valor es un string y contiene la frase clave
        if isinstance(description, str) and 'TOTAL CONSUMOS DE' in description.upper():
            match = total_pattern.search(description)
            if match:
                # Si encuentra el patrón, devuelve solo la parte del texto
                return match.group(1).strip()
        # Si no, devuelve el valor original
        return description

    # <<< CORRECCIÓN >>>: Se aplica la función de limpieza a la columna 'DESCRIPCIÓN'.
    df_txt['DESCRIPCIÓN'] = df_txt['DESCRIPCIÓN'].apply(clean_total_description)


    # 1. Calcular el ancho máximo para cada columna (ahora usando el DataFrame limpio 'df_txt')
    col_widths = {}
    for col in df_txt.columns:
        max_len = max(
            df_txt[col].astype(str).str.len().max(),
            len(col)
        )
        col_widths[col] = int(max_len) + 2

    # Ancho específico para descripción para que no sea tan largo
    if "DESCRIPCIÓN" in col_widths:
         col_widths["DESCRIPCIÓN"] = 50

    with open(output_path, 'w', encoding='utf-8') as f:
        # 2. Escribir el título si existe la fecha
        if latest_date:
            title_month = latest_date.strftime("%B")
            title_text = f"Monthly Consumption {title_month}"
            f.write(title_text.center(sum(col_widths.values())))
            f.write("\n\n")

        # 3. Iterar por cada fila del DataFrame limpio 'df_txt'
        for index, row in df_txt.iterrows():
            line_parts = []
            for col_name in df_txt.columns:
                value = row[col_name]
                
                if isinstance(value, (int, float)):
                    formatted_val = f"{value:,.2f}"
                    line_parts.append(formatted_val.rjust(col_widths[col_name]))
                else:
                    formatted_val = str(value if pd.notna(value) else "")
                    line_parts.append(formatted_val.ljust(col_widths[col_name]))

            line = "".join(line_parts)
            f.write(line.rstrip() + '\n')

    print(f"Data successfully saved to {output_path}")


# This block allows the script to be run directly
if __name__ == "__main__":
    pdf_file_path = "pdfs/04-2025 - Gastos.pdf"  # Make sure this PDF is in the same directory as the script

    if not os.path.exists(pdf_file_path):
        print(f"Error: PDF file not found at '{pdf_file_path}'")
    else:
        print(f"Extracting data from {pdf_file_path}...")
        extracted_data_df, latest_date_found = extract_transactions_from_pdf(pdf_file_path)

        if not extracted_data_df.empty:
            print(f"Extracted {len(extracted_data_df)} records (including headings/totals/blanks).")
            
            # Guardar en Excel (función original)
            save_to_excel(extracted_data_df, latest_date_found)
            
            # <<< NUEVO >>>: Guardar en TXT (nueva función)
            save_to_txt(extracted_data_df, latest_date_found)
            
        else:
            print("No transaction data extracted.")